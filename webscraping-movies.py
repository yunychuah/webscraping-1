
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font


#webpage = 'https://www.boxofficemojo.com/weekend/chart/'
webpage = 'https://www.boxofficemojo.com/year/2023/'

page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

#setting up excel file, creating excel workbook 
wb = xl.Workbook()
#assigning worksheet the ws object which is the active workbook
ws = wb.active
#title 
ws.title = 'Box Office Report'
#headers
ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Release Date'
ws['D1'] = 'Gross'
ws['E1'] = 'Total Gross'
ws['F1'] = '% of Total Gross'

#looking for all the tr tags
movie_rows = soup.findAll('tr')

for x in range(1,6):
    #find td tag in our rows
    td = movie_rows[x].findAll('td')
    #now read from webpage and write to new file
    #number of movie
    no = td[0].text
    #name of movie 
    title = td[1].text
    gross = int(td[5].text.replace(',','').replace('$','')) #5 b/c if u look at inspect on website 
    #it's the fifth one, will highlight it for u
    total_gross = int(td[7].text.replace(',','').replace('$',''))
    release_date = td[8].text
    
    #calculate our percent gross 
    percent_gross = (round(gross/total_gross)*100,2) 
    #round to 2 decimal places

    #done w scraping, now write to excel
    ws['A' + str(x+1)] = no
    ws['B' + str(x+1)] = title
    ws['C' + str(x+1)] = release_date
    ws['D' + str(x+1)] = gross
    ws['E' + str(x+1)] = total_gross
    ws['F' + str(x+1)] = str(percent_gross) + '%'
    #the x + 1: using the x values to iterate through 
    #spreadsheet
    #column a on worksheet

ws.column_dimensions['A'].width = 5


header_font = Font(size=16, bold=True)

for cell in ws[1:1]:
    cell.font = header_font

wb.save('BoxOfficeReport.xlsx')