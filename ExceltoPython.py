import openpyxl as xl 

#open excel document in python using the load_workbook method
wb = xl.load_workbook('example.xlsx')

#prints out sheet names of workbook object
sn = wb.sheetnames

#print out sheet names
print(sn)

sheet1 = wb['Sheet1']
cellA1 = sheet1['A1']

print(cellA1.value) #gets value of cell A1
#pulling out values from row,column,etc.
print(type(cellA1.row)) #type command tells you what type of value it is
#will say <class 'datetime.datetime'>
print(cellA1.column) #gets column 
print(cellA1.coordinate) #gets row and the column

#brings out the word apples 
#the (1,2) tells it to bring out row 1 column 2
print(sheet1.cell(1,2).value)

#gives u number of columns and rows in your doc
print(sheet1.max_row)
print(sheet1.max_column)

#print out all names of fruits using for loop 
for i in range(1, sheet1.max_row+1): #rmbr the +1 or it won't show strawberries
    print(sheet1.cell(i,2).value)#2 b/c u are processing column B 

#get corresponding letters to column number
print(xl.utils.get_column_letter(1))
print(xl.utils.get_column_letter(900))

print(xl.utils.column_index_from_string('AHP'))

for currentrow in sheet1['A1':'C3']:
    print(currentrow)
    for currentcell in currentrow:
        print(currentcell.coordinate, currentcell.value)

#iterates through the columns and rows 
for currentrow in sheet1.iter_rows(min_row=2, max_row=sheet1.max_row, max_col=sheet1.max_column):
    print(currentrow)
    print(currentrow[0].value)
    print(currentrow[1].value)
    print(currentrow[2].value)
